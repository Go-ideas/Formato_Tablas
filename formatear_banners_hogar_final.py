"""
Formatea archivos de banners de Hogar generados en Excel.

Objetivo:
- Partir del Excel crudo/final, sin usar plantilla.
- Generar una copia formateada.
- Mantener textos entre banners sin cuadricula.
- Aplicar cuadricula solo dentro de cada tabla.
- Aplicar bordes gruesos exteriores y separadores gruesos por grupos de columnas.
- No usar guiones bajos ni caracteres como lineas; solo bordes reales de Excel.

Uso desde terminal en VS Code:
    pip install openpyxl
    python formatear_banners_hogar_final.py "BANNERS FINALES  HOGAR 05_05.xlsx"

Opcionalmente puedes indicar archivo de salida:
    python formatear_banners_hogar_final.py "BANNERS FINALES  HOGAR 05_05.xlsx" "BANNERS FORMATEADO.xlsx"
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Callable, Iterable, List, Optional, Tuple

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================================================
# Configuracion visual
# =========================================================

FONT_BASE = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_HEADER = Font(name="Calibri", size=11, bold=False)
FONT_GROUP = Font(name="Calibri", size=11, bold=False)

NO_FILL = PatternFill(fill_type=None)

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
NO_SIDE = Side(style=None)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CENTER_NO_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

QUESTION_RE = re.compile(r"^P\d+", re.IGNORECASE)
SIG_RE = re.compile(r"^[a-zA-Z]+(?:\s+[a-zA-Z]+)*$|^[a-zA-Z]+$")


# =========================================================
# Utilidades
# =========================================================

def is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def is_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def is_percent_like(value) -> bool:
    return is_number(value) and -1 <= float(value) <= 1


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def clean_sig(value) -> str:
    if is_blank(value):
        return ""
    return str(value).strip().replace(" ", "")


def format_percent_with_sig(value, sig: str):
    """Convierte 0.22 + 'bc' en '22%bc'."""
    if not is_percent_like(value):
        return value
    pct = round(float(value) * 100)
    return f"{pct}%{sig}" if sig else value


def copy_border_with(
    border: Border,
    left: Optional[Side] = None,
    right: Optional[Side] = None,
    top: Optional[Side] = None,
    bottom: Optional[Side] = None,
) -> Border:
    """Devuelve un borde nuevo preservando lados no indicados."""
    return Border(
        left=left if left is not None else copy(border.left),
        right=right if right is not None else copy(border.right),
        top=top if top is not None else copy(border.top),
        bottom=bottom if bottom is not None else copy(border.bottom),
    )


def used_max_column(ws) -> int:
    """Ultima columna realmente usada, revisando toda la hoja."""
    max_col = ws.max_column
    while max_col > 1:
        has_value = any(not is_blank(ws.cell(row=r, column=max_col).value) for r in range(1, ws.max_row + 1))
        if has_value:
            return max_col
        max_col -= 1
    return max_col


def row_has_data_from_col(ws, row: int, start_col: int, end_col: int) -> bool:
    return any(not is_blank(ws.cell(row, col).value) for col in range(start_col, end_col + 1))


def normalize_sheet_names(wb) -> None:
    for ws in wb.worksheets:
        name = ws.title.strip().upper().replace("  ", " ")
        if name in {"BANNER2", "BANNER 2"}:
            ws.title = "BANNER 2"
        elif name in {"BANNER1", "BANNER 1"}:
            ws.title = "BANNER 1"


def is_banner_sheet(name: str) -> bool:
    return "BANNER" in name.strip().upper()


def clear_existing_merges(ws) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))


def clear_borders_outside_tables(
    ws,
    row_progress_callback: Optional[Callable[[int, int], None]] = None,
) -> None:
    """Limpia bordes previos para evitar lineas heredadas en textos entre banners."""
    max_row = ws.max_row
    max_col = ws.max_column
    for row in range(1, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).border = Border()
        if row_progress_callback is not None and (row % 80 == 0 or row == max_row):
            row_progress_callback(row, max_row)


# =========================================================
# Preparacion de hoja
# =========================================================

def apply_page_setup(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def apply_column_widths(ws, max_col: int) -> None:
    ws.column_dimensions["A"].width = 41.5
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12.2


def style_text_rows(
    ws,
    max_col: int,
    row_progress_callback: Optional[Callable[[int, int], None]] = None,
) -> None:
    """Estilo para textos de preguntas/notas sin cuadricula."""
    max_row = ws.max_row
    for row in range(1, max_row + 1):
        value = ws.cell(row, 1).value
        c = ws.cell(row, 1)
        c.alignment = ALIGN_LEFT_WRAP if isinstance(value, str) and len(value) > 60 else ALIGN_LEFT
        c.font = FONT_BOLD if isinstance(value, str) and QUESTION_RE.match(value.strip()) else FONT_BASE
        if row_progress_callback is not None and (row % 80 == 0 or row == max_row):
            row_progress_callback(row, max_row)


def clear_borders_only_near_tables(
    ws,
    infos: List[dict],
    max_col: int,
    row_progress_callback: Optional[Callable[[int, int], None]] = None,
) -> None:
    """
    Limpia bordes solo alrededor de las tablas detectadas.
    Evita recorrer toda la hoja cuando hay muchas filas.
    """
    touched_rows = []
    for info in infos:
        start = max(1, info["row_prop"] - 2)
        end = min(ws.max_row, info["row_end"] + 3)
        touched_rows.extend(range(start, end + 1))

    unique_rows = sorted(set(touched_rows))
    total = len(unique_rows)
    for idx, row in enumerate(unique_rows, start=1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).border = Border()
        if row_progress_callback is not None and (idx % 60 == 0 or idx == total):
            row_progress_callback(idx, total)


# =========================================================
# Deteccion de bloques/banner
# =========================================================

def find_header_blocks(ws) -> List[Tuple[int, int, int, int]]:
    """
    Devuelve tuplas:
        (row_proportions, row_group, row_subheader, row_base)

    Estructura esperada aproximada:
        Proportions/Means...
        ** very small base...
        fila grupos
        fila auxiliar/merge
        subencabezados
        letras
        BASE
    """
    blocks: List[Tuple[int, int, int, int]] = []

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        text = clean_text(value).lower()
        if "proportions" not in text or "means" not in text:
            continue

        row_base = None
        for rr in range(row + 1, min(row + 14, ws.max_row + 1)):
            if clean_text(ws.cell(rr, 1).value).upper() == "BASE":
                row_base = rr
                break

        if row_base is None:
            continue

        row_group = row_base - 4
        row_subheader = row_base - 2
        if row_group >= 1 and row_subheader >= 1:
            blocks.append((row, row_group, row_subheader, row_base))

    return blocks


def detect_group_ranges(ws, row_group: int, max_col: int) -> List[Tuple[int, int]]:
    """Detecta rangos de grupos en la fila superior del banner."""
    starts: List[int] = []
    for col in range(2, max_col + 1):
        if not is_blank(ws.cell(row_group, col).value):
            starts.append(col)

    if not starts:
        return [(2, max_col)]

    ranges: List[Tuple[int, int]] = []
    for i, start in enumerate(starts):
        end = starts[i + 1] - 1 if i + 1 < len(starts) else max_col
        ranges.append((start, end))
    return ranges


def find_table_end(ws, row_base: int, next_proportions_row: Optional[int], max_col: int) -> int:
    """
    Encuentra donde termina la tabla.
    Regla: termina en la ultima fila con datos numericos del bloque antes de los textos del siguiente banner.
    """
    hard_end = (next_proportions_row - 1) if next_proportions_row else ws.max_row

    last_data_row = row_base
    for row in range(row_base, hard_end + 1):
        # Si aparece una nueva pregunta antes del siguiente Proportions, ya estamos fuera de tabla.
        label = clean_text(ws.cell(row, 1).value)
        if row > row_base and QUESTION_RE.match(label):
            break

        if row_has_data_from_col(ws, row, 2, max_col):
            last_data_row = row

    # Evita comerse filas de separacion muy largas; la tabla termina en ultima fila con datos.
    return last_data_row


def get_block_infos(ws, max_col: int) -> List[dict]:
    headers = find_header_blocks(ws)
    infos = []
    for idx, (row_prop, row_group, row_subheader, row_base) in enumerate(headers):
        next_prop = headers[idx + 1][0] if idx + 1 < len(headers) else None
        row_end = find_table_end(ws, row_base, next_prop, max_col)
        group_ranges = detect_group_ranges(ws, row_group, max_col)
        group_starts = [start for start, _ in group_ranges]
        infos.append(
            {
                "row_prop": row_prop,
                "row_group": row_group,
                "row_subheader": row_subheader,
                "row_letters": row_subheader + 1,
                "row_base": row_base,
                "row_end": row_end,
                "group_ranges": group_ranges,
                "group_starts": group_starts,
            }
        )
    return infos


# =========================================================
# Significancias
# =========================================================

def append_significance_and_hide_rows(ws, max_col: int, hide_rows: bool = True) -> None:
    """
    En el input, las letras de significancia suelen venir en una fila debajo del porcentaje.
    En el archivo final se pegan al porcentaje: 22%bc.
    Despues se oculta la fila de letras.
    """
    rows_to_hide = []

    for row in range(2, ws.max_row + 1):
        first = ws.cell(row, 1).value
        if not is_blank(first):
            continue

        sig_cols = []
        numeric_above = 0
        for col in range(2, max_col + 1):
            above = ws.cell(row - 1, col).value
            current = clean_sig(ws.cell(row, col).value)

            if is_percent_like(above):
                numeric_above += 1
            if current and SIG_RE.match(current):
                sig_cols.append((col, current))

        # Debe parecer fila de significancias: fila anterior con porcentajes y letras actuales.
        min_numeric = max(2, int((max_col - 1) * 0.25))
        if numeric_above >= min_numeric and sig_cols:
            for col, sig in sig_cols:
                cell = ws.cell(row - 1, col)
                cell.value = format_percent_with_sig(cell.value, sig)
                cell.number_format = "0%"
            rows_to_hide.append(row)

    for row in rows_to_hide:
        if hide_rows:
            ws.row_dimensions[row].hidden = True
            ws.row_dimensions[row].height = 0


# =========================================================
# Bordes jerarquicos
# =========================================================

def internal_border_for_cell(
    row: int,
    col: int,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    group_starts: Iterable[int],
) -> Border:
    """
    Tres niveles:
    - Borde exterior: medio/grueso.
    - Separadores de grupos: medio/grueso.
    - Cuadricula interna: delgada.
    """
    group_starts = set(group_starts)

    left = MEDIUM if col == start_col or col in group_starts else THIN
    right = MEDIUM if col == end_col else THIN
    top = MEDIUM if row == start_row else THIN
    bottom = MEDIUM if row == end_row else THIN

    return Border(left=left, right=right, top=top, bottom=bottom)


def apply_table_grid(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    group_starts: List[int],
    group_ends: Optional[List[int]] = None,
) -> None:
    """Aplica cuadricula completa, incluyendo celdas vacias dentro de la tabla."""
    group_starts_set = set(group_starts)
    group_ends_set = set(group_ends or [])
    top_borders: dict[int, Border] = {}
    mid_borders: dict[int, Border] = {}
    bottom_borders: dict[int, Border] = {}

    for col in range(start_col, end_col + 1):
        left = MEDIUM if col == start_col or col in group_starts_set else THIN
        right = MEDIUM if col == end_col or col in group_ends_set else THIN
        top_borders[col] = Border(left=left, right=right, top=MEDIUM, bottom=THIN)
        mid_borders[col] = Border(left=left, right=right, top=THIN, bottom=THIN)
        bottom_borders[col] = Border(left=left, right=right, top=THIN, bottom=MEDIUM)

    for row in range(start_row, end_row + 1):
        if row == start_row:
            borders = top_borders
        elif row == end_row:
            borders = bottom_borders
        else:
            borders = mid_borders

        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.border = borders[col]

            if col >= 2 and not is_blank(cell.value):
                cell.alignment = ALIGN_CENTER_NO_WRAP
                cell.font = FONT_BASE
                if is_percent_like(cell.value):
                    cell.number_format = "0%"
                elif is_number(cell.value):
                    cell.number_format = "#,##0"


def strengthen_group_right_borders(ws, start_row: int, end_row: int, group_ranges: List[Tuple[int, int]]) -> None:
    """Marca tambien el cierre de cada grupo con linea gruesa."""
    for _, end_col in group_ranges:
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row, end_col)
            cell.border = copy_border_with(cell.border, right=MEDIUM)


def apply_outer_border(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """Refuerza borde exterior sin borrar separadores internos."""
    for col in range(start_col, end_col + 1):
        top_cell = ws.cell(start_row, col)
        bottom_cell = ws.cell(end_row, col)
        top_cell.border = copy_border_with(top_cell.border, top=MEDIUM)
        bottom_cell.border = copy_border_with(bottom_cell.border, bottom=MEDIUM)

    for row in range(start_row, end_row + 1):
        left_cell = ws.cell(row, start_col)
        right_cell = ws.cell(row, end_col)
        left_cell.border = copy_border_with(left_cell.border, left=MEDIUM)
        right_cell.border = copy_border_with(right_cell.border, right=MEDIUM)


# =========================================================
# Encabezados de banner
# =========================================================

def merge_group_headers(ws, row_group: int, group_ranges: List[Tuple[int, int]]) -> None:
    """Une encabezados de grupos con merges reales."""
    for start, end in group_ranges:
        if end >= start:
            ws.merge_cells(start_row=row_group, start_column=start, end_row=row_group + 1, end_column=end)
            cell = ws.cell(row_group, start)
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_GROUP
            cell.fill = NO_FILL


def style_header_block(ws, info: dict, max_col: int) -> None:
    row_group = info["row_group"]
    row_subheader = info["row_subheader"]
    row_letters = info["row_letters"]
    row_base = info["row_base"]
    group_ranges = info["group_ranges"]

    merge_group_headers(ws, row_group, group_ranges)

    # Alturas similares al ejemplo.
    ws.row_dimensions[row_group].height = 45
    ws.row_dimensions[row_group + 1].height = 15
    ws.row_dimensions[row_subheader].height = 48
    ws.row_dimensions[row_letters].height = 18

    # Encabezados y letras.
    for row in range(row_group, row_base + 1):
        for col in range(2, max_col + 1):
            cell = ws.cell(row, col)
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.fill = NO_FILL

    # Refuerza linea inferior de subencabezados y letras.
    for col in range(2, max_col + 1):
        ws.cell(row_subheader, col).border = copy_border_with(ws.cell(row_subheader, col).border, bottom=MEDIUM)
        ws.cell(row_letters, col).border = copy_border_with(ws.cell(row_letters, col).border, bottom=MEDIUM)

    # BASE como numerico.
    for col in range(2, max_col + 1):
        c = ws.cell(row_base, col)
        c.number_format = "#,##0" if is_number(c.value) else c.number_format


# =========================================================
# Filas de datos dentro de tabla
# =========================================================

def style_data_inside_blocks(ws, infos: List[dict], max_col: int) -> None:
    for info in infos:
        row_group = info["row_group"]
        row_end = info["row_end"]
        group_ranges = info["group_ranges"]
        group_starts = info["group_starts"]
        group_ends = [end for _, end in group_ranges]

        apply_table_grid(ws, row_group, row_end, 2, max_col, group_starts, group_ends)

        for row in range(row_group, row_end + 1):
            label_cell = ws.cell(row, 1)
            label_cell.border = Border()  # Columna A queda libre, como en tu ejemplo.
            label_cell.alignment = ALIGN_LEFT
            label_cell.font = FONT_BASE


# =========================================================
# Formateo completo
# =========================================================

def format_sheet(
    ws,
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> None:
    def emit(sheet_progress: float, message: str) -> None:
        if progress_callback is not None:
            progress_callback(max(0.0, min(1.0, sheet_progress)), message)

    # Stage weights to keep progress moving through long operations.
    clear_start = 0.00
    clear_end = 0.02
    text_start = 0.02
    text_end = 0.35
    sig_end = 0.50
    blocks_end = 0.55
    data_start = 0.55
    data_end = 0.90
    headers_start = 0.90
    headers_end = 0.98
    final_end = 1.00

    emit(clear_start, 'Preparando hoja')
    clear_existing_merges(ws)
    max_col = used_max_column(ws)
    emit(clear_end, 'Preparacion completada')

    emit(text_start, 'Aplicando configuracion visual base')
    apply_page_setup(ws)
    apply_column_widths(ws, max_col)
    style_text_rows(
        ws,
        max_col,
        row_progress_callback=lambda row, total: emit(
            text_start + ((text_end - text_start) * (row / max(1, total))),
            f'Aplicando estilo de texto ({row}/{total})',
        ),
    )

    emit(text_end, 'Integrando significancias')
    append_significance_and_hide_rows(ws, max_col, hide_rows=True)
    emit(sig_end, 'Significancias integradas')

    max_col = used_max_column(ws)
    infos = get_block_infos(ws, max_col)
    emit(blocks_end, 'Bloques detectados')

    # Apply data grid block by block to report useful progress.
    total_blocks = len(infos)
    if total_blocks == 0:
        emit(data_end, 'No se detectaron bloques para formatear')
    else:
        for idx, info in enumerate(infos, start=1):
            style_data_inside_blocks(ws, [info], max_col)
            emit(
                data_start + ((data_end - data_start) * (idx / total_blocks)),
                f'Aplicando cuadricula y formato ({idx}/{total_blocks})',
            )

    # Apply headers in a separate pass.
    if total_blocks == 0:
        emit(headers_end, 'Sin encabezados para aplicar')
    else:
        for idx, info in enumerate(infos, start=1):
            style_header_block(ws, info, max_col)
            emit(
                headers_start + ((headers_end - headers_start) * (idx / total_blocks)),
                f'Aplicando encabezados ({idx}/{total_blocks})',
            )

    emit(headers_end, 'Ajustes finales')
    emit(final_end, 'Hoja completada')


def format_workbook(
    input_path: str | Path,
    output_path: str | Path,
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> None:
    input_path = Path(input_path)
    output_path = Path(output_path)

    if not input_path.exists():
        raise FileNotFoundError(f'No existe el archivo de entrada: {input_path}')

    def emit(progress: float, message: str) -> None:
        if progress_callback is not None:
            progress_callback(max(0.0, min(1.0, progress)), message)

    emit(0.01, 'Cargando archivo Excel')
    wb = load_workbook(input_path)
    normalize_sheet_names(wb)

    banner_sheets = [ws for ws in wb.worksheets if is_banner_sheet(ws.title)]
    emit(0.05, f'Archivo cargado. Hojas objetivo: {len(banner_sheets)}')

    if not banner_sheets:
        emit(0.90, 'No se encontraron hojas BANNER. Guardando salida')
        wb.save(output_path)
        emit(1.00, 'Proceso completado')
        print(f'Archivo generado: {output_path}')
        return

    process_start = 0.05
    process_end = 0.95
    process_span = process_end - process_start

    for index, ws in enumerate(banner_sheets, start=1):
        print(f'Formateando hoja: {ws.title}')

        sheet_start = process_start + process_span * ((index - 1) / len(banner_sheets))
        sheet_end = process_start + process_span * (index / len(banner_sheets))
        sheet_span = sheet_end - sheet_start

        def on_sheet_progress(sheet_progress: float, step_message: str) -> None:
            emit(
                sheet_start + (sheet_span * sheet_progress),
                f'Hoja {index}/{len(banner_sheets)} ({ws.title}): {step_message}',
            )

        format_sheet(ws, progress_callback=on_sheet_progress)

    emit(0.97, 'Guardando archivo de salida')
    wb.save(output_path)
    emit(1.00, 'Proceso completado')
    print(f'Archivo generado: {output_path}')

def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python formatear_banners_hogar_final.py <archivo_input.xlsx> [archivo_salida.xlsx]")
        sys.exit(1)

    input_path = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2])
    else:
        output_path = input_path.with_name(input_path.stem + " FORMATEADO.xlsx")

    format_workbook(input_path, output_path)


if __name__ == "__main__":
    main()

